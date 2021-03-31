#+ 数据科学常用工具
import matplotlib as mpl
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.style as style
import seaborn as sns
from sklearn.preprocessing import PowerTransformer
import category_encoders as ce
from sklearn.model_selection import StratifiedKFold, KFold
from joblib import Parallel, delayed
import multiprocessing
from scipy import stats
from pandas.api.types import is_datetime64_any_dtype as is_datetime
from pandas.api.types import is_categorical_dtype
from sklearn.model_selection import KFold
from sklearn.model_selection._split import _BaseKFold, indexable, _num_samples
from sklearn.utils.validation import _deprecate_positional_args
from openpyxl import load_workbook, Workbook
import xlrd


# ---------------------------------
# 可视化工具
# ---------------------------------
def set_format():
    pd.set_eng_float_format(accuracy=2, use_eng_prefix=False)


def set_matplotlib():
    plt.rcParams['font.sans-serif'] = ['Arial Unicode MS']
    mpl.rcParams['figure.dpi'] = 100
    # style.available
    # style.use('ggplot')


def whiteboard(row=1, col=1, dpi=100):
    fig, ax = plt.subplots(row, col, figsize=(6.4, 4.8), dpi=dpi)
    return (fig, ax)


def count_plot(df, var, type='bar'):
    fig, ax = whiteboard()
    counts = df[var].value_counts()
    if type == 'bar':
        counts.head(15).sort_values().plot.barh(ax=ax)
    else:
        counts.sort_index().plot.line(ax=ax)
    ax.set_xlabel('count')
    ax.set_ylabel('value')
    ax.set_title(var)


def kde_by_target(df_raw, var, target, cut=0.99):
    if cut is not None:
        upper = df_raw[var].quantile(cut)
        df = df_raw[df_raw[var] <= upper]
    else:
        df = df_raw
    # 大多数的变量都有很长的尾部，为了看得清楚做截尾
    fig, ax = whiteboard()
    for y in df[target].unique():
        sub = df[df[target] == y]
        sns.distplot(sub[var], hist=False, ax=ax, label=str(y),
                     kde_kws={"lw": 0.7})
    ax.legend()


def series_plot(series, xdate=True, xlabel='date', ylabel='', title=''):
    fig, ax = whiteboard()
    series.plot(ax=ax, linewidth=1.0)
    if xdate:
        fig.autofmt_xdate()

    ax.set_ylabel(ylabel)
    ax.set_xlabel(xlabel)
    ax.set_title(title)


def sns_sparse_xticks(plot_, freq):
    for ind, label in enumerate(plot_.get_xticklabels()):
        if ind % freq == 0:  # every 10th label is kept
            label.set_visible(True)
        else:
            label.set_visible(False)


# ---------------------------------
# 描述统计
# ---------------------------------
def count_na(df):
    missing_dict = df.isna().sum()
    return missing_dict[missing_dict > 0]


def groupby_info(df, col, target):
    return df.groupby(target)[col].agg(mean=np.mean, median=np.median, mode=lambda x: stats.mode(x)[0][0], max=np.max,
                                       min=np.min, std=np.std)


# ---------------------------------
# 特征清洗
# ---------------------------------
# 缺失值
def replace(x, from_, to_):
    tmp = x.copy()
    tmp[tmp == from_] = to_
    return tmp


def group_fillna(df, col, target, method='mean'):
    if method == 'mean':
        result = df.groupby([target], sort=False)[col].apply(lambda x: x.fillna(x.mean()))
    elif method == 'median':
        result = df.groupby([target], sort=False)[col].apply(lambda x: x.fillna(x.median()))
    return result


# 异常值
def windsorize(series, upper, lower):
    return series.clip(lower=lower, upper=upper)


def cap(x, extreme=5):
    # 生成分位数
    width = (x.quantile(0.75) - x.quantile(0.25)) / 2
    median = x.median()
    return x.clip(median - extreme * width, median + extreme * width)


# ---------------------------------
# 单特征变换
# ---------------------------------
def box_cox(x_train, x_test=None):
    bc = PowerTransformer(method='box-cox')
    bc = bc.fit(x_train)
    x_train_bc = bc.transform(x_train)
    if x_test is not None:
        x_test_bc = bc.transform(x_test)
    else:
        x_test_bc = None
    return (x_train_bc, x_test_bc)


def standardize(x_train, x_test=None, cut=None):
    """
    cut: 截断cut倍标准差
    """
    avg, var = x_train.mean(), x_train.std()
    x_train_s = (x_train - avg) / var
    if cut is not None:
        x_train_s = windsorize(x_train_s, cut, -cut)
    if x_test is not None:
        x_test_s = (x_test - avg) / var
        if cut is not None:
            x_test_s = windsorize(x_test_s, cut, -cut)
    else:
        x_test_s = None
    return (x_train_s, x_test_s)


def bin(x, n_scatter=10):
    """
    连续变量分箱
    """
    result = pd.qcut(x, n_scatter)
    return result


# ---------------------------------
# 特征编码
# ---------------------------------
# One-hot encoding for categorical columns with get_dummies
def one_hot_encoder(df, categorical_columns=None, nan_as_category=True, min_count=100, inplace=True):
    original_columns = list(df.columns)
    if categorical_columns is None:  # 自动查找离散变量
        categorical_columns = [col for col in df.columns if df[col].dtype == 'object']
    result = pd.get_dummies(df, columns=categorical_columns, dummy_na=nan_as_category)
    new_columns = [c for c in result.columns if c not in original_columns]
    cat_columns = [c for c in original_columns if c not in result.columns]
    if not inplace:  # 返回值包含原始离散特征
        for c in cat_columns:
            result[c] = df[c]
    i = 0
    for c in new_columns:  # 合并稀少的类
        if (result[c].sum() < min_count) or ((result.shape[0] - result[c].sum()) < min_count):
            i += 1
            del result[c]
            new_columns.remove(c)
    if i == 0:
        del result[c]  # 哑变量
    return result, new_columns


# 连续特征离散化
def one_hot_encoder_continus(df, col, n_scatter=10, nan_as_category=True, min_count=100):
    df[col + '_scatter'] = pd.qcut(df[col], n_scatter)
    result, new_cols = one_hot_encoder(df, [col + '_scatter'], nan_as_category=nan_as_category, min_count=min_count,
                                       inplace=True)
    return result, new_cols


# count encoding
def count_encoding(li):
    temp = pd.Series(li)
    result = temp.map(temp.value_counts())
    return result


# cv method for the following encoders
def cv_encoding(encoding_func, X, y, cols, target_type=None, n_splits=10, **kwargs):
    if target_type is None:
        if y.dtype == int or y.dtype == object:
            target_type = 'cat'
        else:
            target_type = 'con'

    if target_type == 'cat':
        kf = StratifiedKFold(n_splits=n_splits)
        split = kf.split(X, y)
    else:
        kf = KFold(n_splits=n_splits)
        split = kf.split(X)

    collect = []
    for train_index, test_index in split:
        X_train, X_test = X.loc[train_index], X.loc[test_index]
        y_train, y_test = y.loc[train_index], y.loc[test_index]
        collect.append(encoding_func(X_train, y_train, cols, X_test, **kwargs))

    return pd.concat(collect)


# target encoding
def target_encoding(X_fit, y_fit, cols, X_test=None, smoothing=0):
    """
    针对continuous和binomial target做均值编码
    X_fit: 用来计算encoding的df, 包含cols
    y_fit: encoding的target
    X_test: 需要transform的对象
    cols: 需要encoding的列
    smoothing: prior权重
    """
    if X_test is None:
        X_test = X_fit
    encoder = ce.TargetEncoder(cols=cols, smoothing=smoothing)
    encoder.fit(X_fit, y_fit)
    result = encoder.transform(X_test)
    return result


# WOE encoding
def woe_encoding(X_fit, y_fit, cols, X_test=None, sigma=0):
    """
    只针对binomial target
    X_fit: 用来计算encoding的df, 包含cols
    y_fit: encoding的target
    X_test: 需要transform的对象
    cols: 需要encoding的列
    sigma: 添加噪声的标准差，防止过拟合
    """
    if X_test is None:
        X_test = X_fit
    encoder = ce.WOEEncoder(cols=cols, sigma=sigma)
    encoder.fit(X_fit, y_fit)
    result = encoder.transform(X_test)
    return result


# James-Stein encoding
def js_encoding(X_fit, y_fit, cols, X_test=None, model='independent'):
    """
    只针对continuous target
    X_fit: 用来计算encoding的df, 包含cols
    y_fit: encoding的target
    X_test: 需要transform的对象
    cols: 需要encoding的列
    model: 'pooled' or 'independent'；pooled是假设所有个体具有相同的方差，与casi书中定义一致
    """
    if X_test is None:
        X_test = X_fit
    encoder = ce.JamesSteinEncoder(cols=cols, model=model)
    encoder.fit(X_fit, y_fit)
    result = encoder.transform(X_test)
    return result


# ---------------------------------
# Time Series Analysis
# ---------------------------------

# kalman filter rolling least square
def RLS(x, y, beta_init=None, R_init=None, delta=0.02, Ve=0.001, intercept=True):
    n, p = x.shape
    
    if intercept:
        intercept_ = np.ones([n, 1])
        x = np.hstack([intercept_, x])
        p += 1
    
    yhat = np.zeros(n)
    e = np.zeros(n)
    Q = np.zeros(n)
    
    if R_init is None:
        R = np.zeros([p, p])
    else:
        R = R_init
    beta = np.zeros([p, n])
    
    Vw = delta / (1 - delta) * np.eye(p)
    # Ve = 
    
    # initialize
    if beta_init is not None:
        beta[:, 0] = beta_init
    
    # kalman loop
    for t in range(n):
        if t > 0:
            beta[:, t] = beta[:, t-1]  # state prediction
            R = P + Vw  # state covariance prediction
        xt = x[t, :]
        yhat[t] = xt.dot(beta[:, t])  # measurement prediction
        Q[t] = xt.dot(R).dot(xt.T) + Ve  # measurement variance
        
        e[t] = y[t] - yhat[t]  # measurement residual
        K = R.dot(xt) / Q[t]  # kalman gain
        
        beta[:, t] = beta[:, t] + K * e[t]  # state update
        P = (1 - K.dot(xt)) * R
    
    return beta


# kalman filter rls with ols start
def ols_start_rls(x, y, start=100, delta=0.1, Ve=0.002, intercept=True):
    x_start = x[:start, :]
    y_start = y[:start]
    
    if intercept:
        intercept_ = np.ones([start, 1])
        x_start = np.hstack([intercept_, x_start])
    _, p = x_start.shape
    beta_init = pinv(x_start.T.dot(x_start)).dot(x_start.T.dot(y_start))
    e = y_start - x_start.dot(beta_init)
    sig_hat_square = e.dot(e) / (start - p - 1)
    R_init = sig_hat_square * pinv(x_start.T.dot(x_start))
    
    x_tail = x[start:, :]
    y_tail = y[start:]
    
    beta_tail = RLS(x_tail, y_tail, beta_init=beta_init, R_init=R_init, delta=delta, Ve=Ve, intercept=intercept)
    return (beta_init, beta_tail)


# ---------------------------------
# 其他
# ---------------------------------

# 查表名
def get_sheetname(fp):
    xls = xlrd.open_workbook(fp, on_demand=True)
    return xls.sheet_names()


def _mdf_func(func, name, group):
    return func(group), name


# 并行df.groupby(col).apply
def applyParallel(dfGrouped, func):
    retLst, top_index = zip(*Parallel(n_jobs=multiprocessing.cpu_count())(
        delayed(_mdf_func)(func, name, group) for name, group in dfGrouped))
    res = pd.concat(retLst, axis=1).T
    res.index = top_index
    return res


# 写入结果到excel
def append_df_to_excel(df, fp, sheet_name, index=False):
    try:
        book = load_workbook(fp)
    except:
        wb = Workbook()
        wb.save(fp)
        book = load_workbook(fp)
        print("creating new file...")

    writer = pd.ExcelWriter(fp, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, sheet_name=sheet_name, index=index)
    writer.save()
    print('done;')


def reduce_mem_usage(df, use_float16=False):
    """
    Iterate through all the columns of a dataframe and modify the data type to reduce memory usage.
    """

    start_mem = df.memory_usage().sum() / 1024 ** 2
    print("Memory usage of dataframe is {:.2f} MB".format(start_mem))

    for col in df.columns:
        if is_datetime(df[col]) or is_categorical_dtype(df[col]):
            continue
        col_type = df[col].dtype

        if col_type != object:
            c_min = df[col].min()
            c_max = df[col].max()
            if str(col_type)[:3] == "int":
                if c_min > np.iinfo(np.int8).min and c_max < np.iinfo(np.int8).max:
                    df[col] = df[col].astype(np.int8)
                elif c_min > np.iinfo(np.int16).min and c_max < np.iinfo(np.int16).max:
                    df[col] = df[col].astype(np.int16)
                elif c_min > np.iinfo(np.int32).min and c_max < np.iinfo(np.int32).max:
                    df[col] = df[col].astype(np.int32)
                elif c_min > np.iinfo(np.int64).min and c_max < np.iinfo(np.int64).max:
                    df[col] = df[col].astype(np.int64)
            else:
                if use_float16 and c_min > np.finfo(np.float16).min and c_max < np.finfo(np.float16).max:
                    df[col] = df[col].astype(np.float16)
                elif c_min > np.finfo(np.float32).min and c_max < np.finfo(np.float32).max:
                    df[col] = df[col].astype(np.float32)
                else:
                    df[col] = df[col].astype(np.float64)
        else:
            df[col] = df[col].astype("category")

    end_mem = df.memory_usage().sum() / 1024 ** 2
    print("Memory usage after optimization is: {:.2f} MB".format(end_mem))
    print("Decreased by {:.1f}%".format(100 * (start_mem - end_mem) / start_mem))

    return df


# group cv
class PurgedGroupTimeSeriesSplitStacking(_BaseKFold):
    """Time Series cross-validator variant with non-overlapping groups.
    Allows for a gap in groups to avoid potentially leaking info from
    train into test if the model has windowed or lag features.
    Provides train/test indices to split time series data samples
    that are observed at fixed time intervals according to a
    third-party provided group.
    In each split, test indices must be higher than before, and thus shuffling
    in cross validator is inappropriate.
    This cross-validation object is a variation of :class:`KFold`.
    In the kth split, it returns first k folds as train set and the
    (k+1)th fold as test set.
    The same group will not appear in two different folds (the number of
    distinct groups has to be at least equal to the number of folds).
    Note that unlike standard cross-validation methods, successive
    training sets are supersets of those that come before them.
    Read more in the :ref:`User Guide <cross_validation>`.

    Parameters
    ----------
    n_splits : int, default=5
        Number of splits. Must be at least 2.
    stacking_mode : bool, default=True
        Whether to provide an additional set to test a stacking classifier or not.
    max_train_group_size : int, default=Inf
        Maximum group size for a single training set.
    max_val_group_size : int, default=Inf
        Maximum group size for a single validation set.
    max_test_group_size : int, default=Inf
        We discard this number of groups from the end of each train split, if stacking_mode = True and None
        it defaults to max_val_group_size.
    val_group_gap : int, default=None
        Gap between train and validation
    test_group_gap : int, default=None
        Gap between validation and test, if stacking_mode = True and None
        it defaults to val_group_gap.

    Example
    ----------
    cv = PurgedGroupTimeSeriesSplitStacking(
        n_splits=5,
        stacking_mode=False,
        max_train_group_size=100,
        val_group_gap=20
    )
    """

    @_deprecate_positional_args
    def __init__(self,
                 n_splits=5,
                 *,
                 stacking_mode=True,
                 max_train_group_size=np.inf,
                 max_val_group_size=np.inf,
                 max_test_group_size=np.inf,
                 val_group_gap=None,
                 test_group_gap=None,
                 verbose=False
                 ):
        super().__init__(n_splits, shuffle=False, random_state=None)
        self.max_train_group_size = max_train_group_size
        self.max_val_group_size = max_val_group_size
        self.max_test_group_size = max_test_group_size
        self.val_group_gap = val_group_gap
        self.test_group_gap = test_group_gap
        self.verbose = verbose
        self.stacking_mode = stacking_mode

    def split(self, X, y=None, groups=None):
        if self.stacking_mode:
            return self.split_ensemble(X, y, groups)
        else:
            return self.split_standard(X, y, groups)

    def split_standard(self, X, y=None, groups=None):
        """Generate indices to split data into training and validation set.

        Parameters
        ----------
        X : array-like of shape (n_samples, n_features)
            Training data, where n_samples is the number of samples
            and n_features is the number of features.
        y : array-like of shape (n_samples,)
            Always ignored, exists for compatibility.
        groups : array-like of shape (n_samples,)
            Group labels for the samples used while splitting the dataset into
            train/validation set.
        Yields
        ------
        train : ndarray
            The training set indices for that split.
        val : ndarray
            The validation set indices for that split.
        """
        if groups is None:
            raise ValueError(
                "The 'groups' parameter should not be None")
        X, y, groups = indexable(X, y, groups)
        n_splits = self.n_splits
        group_gap = self.val_group_gap
        max_val_group_size = self.max_val_group_size
        max_train_group_size = self.max_train_group_size
        n_folds = n_splits + 1
        group_dict = {}
        u, ind = np.unique(groups, return_index=True)
        unique_groups = u[np.argsort(ind)]
        n_samples = _num_samples(X)
        n_groups = _num_samples(unique_groups)
        for idx in np.arange(n_samples):
            if (groups[idx] in group_dict):
                group_dict[groups[idx]].append(idx)
            else:
                group_dict[groups[idx]] = [idx]
        if n_folds > n_groups:
            raise ValueError(
                ("Cannot have number of folds={0} greater than"
                 " the number of groups={1}").format(n_folds,
                                                     n_groups))

        group_val_size = min(n_groups // n_folds, max_val_group_size)
        group_val_starts = range(n_groups - n_splits * group_val_size,
                                 n_groups, group_val_size)
        for group_val_start in group_val_starts:
            train_array = []
            val_array = []

            group_st = max(0, group_val_start - group_gap - max_train_group_size)
            for train_group_idx in unique_groups[group_st:(group_val_start - group_gap)]:
                train_array_tmp = group_dict[train_group_idx]

                train_array = np.sort(np.unique(
                    np.concatenate((train_array,
                                    train_array_tmp)),
                    axis=None), axis=None)

            train_end = train_array.size

            for val_group_idx in unique_groups[group_val_start:
            group_val_start +
            group_val_size]:
                val_array_tmp = group_dict[val_group_idx]
                val_array = np.sort(np.unique(
                    np.concatenate((val_array,
                                    val_array_tmp)),
                    axis=None), axis=None)

            val_array = val_array[group_gap:]

            if self.verbose > 0:
                pass

            yield [int(i) for i in train_array], [int(i) for i in val_array]

    def split_ensemble(self, X, y=None, groups=None):
        """Generate indices to split data into training, validation and test set.
        Parameters
        ----------
        X : array-like of shape (n_samples, n_features)
            Training data, where n_samples is the number of samples
            and n_features is the number of features.
        y : array-like of shape (n_samples,)
            Always ignored, exists for compatibility.
        groups : array-like of shape (n_samples,)
            Group labels for the samples used while splitting the dataset into
            train/test set.
        Yields
        ------
        train : ndarray
            The training set indices for that split.
        val : ndarray
            The validation set indices for that split (testing indices for base classifiers).
        test : ndarray
            The testing set indices for that split (testing indices for final classifier)
        """

        if groups is None:
            raise ValueError(
                "The 'groups' parameter should not be None")

        X, y, groups = indexable(X, y, groups)
        n_splits = self.n_splits
        val_group_gap = self.val_group_gap
        test_group_gap = self.test_group_gap
        if test_group_gap is None:
            test_group_gap = val_group_gap
        max_train_group_size = self.max_train_group_size
        max_val_group_size = self.max_val_group_size
        max_test_group_size = self.max_test_group_size
        if max_test_group_size is None:
            max_test_group_size = max_val_group_size

        n_folds = n_splits + 1
        group_dict = {}
        u, ind = np.unique(groups, return_index=True)
        unique_groups = u[np.argsort(ind)]
        n_samples = _num_samples(X)
        n_groups = _num_samples(unique_groups)

        for idx in np.arange(n_samples):
            if (groups[idx] in group_dict):
                group_dict[groups[idx]].append(idx)
            else:
                group_dict[groups[idx]] = [idx]
        if n_folds > n_groups:
            raise ValueError(
                ("Cannot have number of folds={0} greater than"
                 " the number of groups={1}").format(n_folds,
                                                     n_groups))

        group_val_size = min(n_groups // n_folds, max_val_group_size)
        group_test_size = min(n_groups // n_folds, max_test_group_size)

        group_test_starts = range(n_groups - n_splits * group_test_size, n_groups, group_test_size)
        train_indices = []
        val_indices = []
        test_indices = []

        for group_test_start in group_test_starts:

            train_array = []
            val_array = []
            test_array = []

            val_group_st = max(max_train_group_size + val_group_gap,
                               group_test_start - test_group_gap - max_val_group_size)

            train_group_st = max(0, val_group_st - val_group_gap - max_train_group_size)

            for train_group_idx in unique_groups[train_group_st:(val_group_st - val_group_gap)]:
                train_array_tmp = group_dict[train_group_idx]

                train_array = np.sort(np.unique(
                    np.concatenate((train_array,
                                    train_array_tmp)),
                    axis=None), axis=None)

            train_end = train_array.size

            for val_group_idx in unique_groups[val_group_st:(group_test_start - test_group_gap)]:
                val_array_tmp = group_dict[val_group_idx]
                val_array = np.sort(np.unique(
                    np.concatenate((val_array,
                                    val_array_tmp)),
                    axis=None), axis=None)

            val_array = val_array[val_group_gap:]

            for test_group_idx in unique_groups[group_test_start:(group_test_start + group_test_size)]:
                test_array_tmp = group_dict[test_group_idx]
                test_array = np.sort(np.unique(
                    np.concatenate((test_array,
                                    test_array_tmp)),
                    axis=None), axis=None)

            test_array = test_array[test_group_gap:]

            yield [int(i) for i in train_array], [int(i) for i in val_array], [int(i) for i in test_array]





